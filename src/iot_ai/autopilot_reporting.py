# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Required Notice: Copyright 2026 IoT-AI.Tech / Dr.-Ing. Babak Sorkhpour
# Author: Dr.-Ing. Babak Sorkhpour, with AI assistance
# Version: 6.7.0-beta.5 | Date: 2026-08-08
"""Deterministic final reporting for autonomous Task/Meeting/Multi-Coder runs."""
from __future__ import annotations

import csv
import hashlib
import json
import tempfile
from pathlib import Path
from typing import Any

from .paths import state_root
from .util import atomic_json, atomic_text, sha256_file, utc_now

TASK_COLUMNS = (
    "task_id", "title", "backend", "authority_basis", "priority", "initial_state",
    "acceptance", "meeting", "multi_coder", "tests", "repairs", "iterations",
    "final_state", "remaining_work", "blocker_next_actor", "next_action", "evidence",
)
PROVIDER_COLUMNS = (
    "task_id", "iteration", "seat", "provider", "model_requested", "model_served",
    "status", "substantive", "failure_class", "decision",
)
ITERATION_COLUMNS = (
    "task_id", "iteration", "stage", "decision", "failure_fingerprint", "new_evidence",
    "meeting_id", "run_id", "started_at", "finished_at",
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return str(value)


def _task_row(value: dict[str, Any]) -> dict[str, str]:
    ac = value.get("acceptance") or {}
    ac_text = f"{ac.get('passed', 0)}/{ac.get('total', 0)}"
    if ac.get("basis"):
        ac_text += f" {ac['basis']}"
    declared_total = ac.get("declared_total")
    if declared_total is not None and int(declared_total or 0) != int(ac.get("total") or 0):
        ac_text += f"; declared={declared_total}"
    if ac.get("status"):
        ac_text += f" ({ac['status']})"
    meeting = value.get("meeting") or {}
    multi = value.get("multi_coder") or {}
    tests = value.get("tests") or {}
    evidence = value.get("evidence") or []
    final_state = _text(value.get("final_state"))
    blocker = _text(value.get("blocker_next_actor"))
    if final_state == "COMPLETE":
        remaining_work, next_action = "none", "none"
    elif final_state == "TECHNICAL_COMPLETE_AWAITING_FOUNDER":
        remaining_work, next_action = "technical work complete; final human decision pending", "founder accept, reject, or rework"
    elif final_state == "BUDGET_EXHAUSTED":
        remaining_work, next_action = "bounded run budget exhausted; checkpoint preserved", "resume the same natural-language goal from checkpoint"
    elif final_state in {"EXTERNALLY_BLOCKED", "AUTHORITY_BLOCKED", "SAFETY_BLOCKED"}:
        remaining_work, next_action = blocker or "external or authority blocker", blocker or "resolve blocker and resume"
    else:
        remaining_work, next_action = blocker or "unresolved acceptance evidence or verification", blocker or "resume closed loop"
    return {
        "task_id": _text(value.get("task_id")),
        "title": _text(value.get("title")),
        "backend": _text(value.get("backend")),
        "authority_basis": _text(value.get("authority_basis")),
        "priority": _text(value.get("priority")),
        "initial_state": _text(value.get("initial_state")),
        "acceptance": ac_text,
        "meeting": _text(meeting.get("decision") or meeting.get("status") or "not-run"),
        "multi_coder": _text(multi.get("decision") or multi.get("status") or "not-run"),
        "tests": _text(tests.get("summary") or tests.get("decision") or "not-run"),
        "repairs": _text(value.get("repairs", 0)),
        "iterations": _text(value.get("iteration_count", 0)),
        "final_state": final_state,
        "remaining_work": remaining_work,
        "blocker_next_actor": blocker,
        "next_action": next_action,
        "evidence": "; ".join(_text(item) for item in evidence if item),
    }


def build_report(
    *,
    intent: dict[str, Any],
    backend_receipt: dict[str, Any] | None,
    tasks: list[dict[str, Any]],
    providers: list[dict[str, Any]],
    iterations: list[dict[str, Any]],
    human_decisions: list[dict[str, Any]],
    terminal_state: str,
    blockers: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    iteration_counts: dict[str, int] = {}
    for row in iterations:
        task_id = str(row.get("task_id") or "")
        if task_id and task_id != "*":
            iteration_counts[task_id] = iteration_counts.get(task_id, 0) + 1
    enriched_tasks = [
        {**item, "iteration_count": iteration_counts.get(str(item.get("task_id") or ""), 0)}
        for item in tasks
    ]
    task_rows = [_task_row(item) for item in enriched_tasks]
    substantive_provider_rows = sum(1 for item in providers if item.get("substantive"))
    failed_provider_rows = sum(1 for item in providers if str(item.get("status") or "") != "pass")
    exact_model_rows = sum(1 for item in providers if item.get("model_served"))
    report: dict[str, Any] = {
        "schema": "iot-ai.autonomous-run-report.v2",
        "generated_at": utc_now(),
        "intent": intent,
        "backend_receipt": backend_receipt or {},
        "terminal_state": terminal_state,
        "summary": {
            "task_count": len(task_rows),
            "complete": sum(1 for item in tasks if item.get("final_state") == "COMPLETE"),
            "awaiting_founder": sum(1 for item in tasks if item.get("final_state") == "TECHNICAL_COMPLETE_AWAITING_FOUNDER"),
            "blocked": sum(1 for item in tasks if "BLOCKED" in str(item.get("final_state") or "")),
            "needs_work": sum(1 for item in tasks if item.get("final_state") == "NEEDS_WORK"),
            "provider_rows": len(providers),
            "provider_substantive_rows": substantive_provider_rows,
            "provider_failed_rows": failed_provider_rows,
            "provider_exact_model_rows": exact_model_rows,
            "iterations": len(iterations),
        },
        "task_table": task_rows,
        "provider_table": providers,
        "iteration_table": iterations,
        "human_decisions": human_decisions,
        "blockers": blockers or [],
        "production_claim": False,
    }
    report["digest"] = hashlib.sha256(
        json.dumps(report, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")
    ).hexdigest()
    return report


def _markdown(report: dict[str, Any]) -> str:
    lines = [
        "# MC-GPT Autonomous Run Report",
        "",
        f"- Generated: `{report['generated_at']}`",
        f"- Terminal state: **{report['terminal_state']}**",
        f"- Intent digest: `{(report.get('intent') or {}).get('digest', '')}`",
        f"- Report digest: `{report.get('digest', '')}`",
        "- Production claim: `false`",
        "",
        "## Task table",
        "",
        "| Task ID | Title | Backend | Authority | Priority | Initial | AC Pass/Total | Meeting | Multi-Coder | Tests | Repairs | Iterations | Final State | Remaining Work | Next Actor | Next Action | Evidence |",
        "|---|---|---|---|---|---|---|---|---|---|---:|---:|---|---|---|---|---|",
    ]
    for row in report.get("task_table", []):
        values = [row.get(name, "") for name in TASK_COLUMNS]
        lines.append("| " + " | ".join(str(value).replace("|", "\\|").replace("\n", " ") for value in values) + " |")
    lines.extend(["", "## Provider participation", "", "| Task | Iteration | Seat | Provider | Requested | Served | Status | Substantive | Failure | Decision |", "|---|---:|---|---|---|---|---|---|---|---|"])
    for row in report.get("provider_table", []):
        values = [row.get(name, "") for name in PROVIDER_COLUMNS]
        lines.append("| " + " | ".join(str(value).replace("|", "\\|").replace("\n", " ") for value in values) + " |")
    lines.extend(["", "## Iterations", "", "| Task | # | Stage | Decision | Failure Fingerprint | New Evidence | Meeting | Run | Started | Finished |", "|---|---:|---|---|---|---|---|---|---|---|"])
    for row in report.get("iteration_table", []):
        values = [row.get(name, "") for name in ITERATION_COLUMNS]
        lines.append("| " + " | ".join(str(value).replace("|", "\\|").replace("\n", " ") for value in values) + " |")
    if report.get("blockers"):
        lines.extend(["", "## Remaining blockers", ""])
        for blocker in report["blockers"]:
            lines.append(f"- `{blocker.get('code', 'blocker')}` — {blocker.get('summary', '')} · next actor: `{blocker.get('next_actor', 'unknown')}`")
    return "\n".join(lines) + "\n"


def write_report_bundle(user_home: Path, report: dict[str, Any], output: Path | None = None) -> dict[str, Any]:
    root = output or (state_root(user_home) / "autopilot" / "reports" / f"run-{report['digest'][:16]}")
    root.mkdir(parents=True, exist_ok=True)
    json_path = root / "AUTONOMOUS_RUN_REPORT.json"
    md_path = root / "AUTONOMOUS_RUN_REPORT.md"
    tasks_csv = root / "TASKS.csv"
    providers_csv = root / "PROVIDERS.csv"
    iterations_csv = root / "ITERATIONS.csv"
    xlsx_path = root / "AUTONOMOUS_RUN_REPORT.xlsx"
    atomic_json(json_path, report)
    atomic_text(md_path, _markdown(report), 0o600)

    def write_csv(path: Path, columns: tuple[str, ...], rows: list[dict[str, Any]]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(columns), extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({key: _text(row.get(key)) for key in columns})

    write_csv(tasks_csv, TASK_COLUMNS, list(report.get("task_table", [])))
    write_csv(providers_csv, PROVIDER_COLUMNS, list(report.get("provider_table", [])))
    write_csv(iterations_csv, ITERATION_COLUMNS, list(report.get("iteration_table", [])))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        wb.remove(wb.active)
        for title, columns, values in (
            ("Tasks", TASK_COLUMNS, report.get("task_table", [])),
            ("Providers", PROVIDER_COLUMNS, report.get("provider_table", [])),
            ("Iterations", ITERATION_COLUMNS, report.get("iteration_table", [])),
        ):
            sheet = wb.create_sheet(title)
            sheet.append(list(columns))
            for value in values:
                sheet.append([_text(value.get(column)) for column in columns])
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for index, column in enumerate(columns, start=1):
                width = min(40, max(12, len(column) + 2, *(len(str(row.get(column, ""))) + 2 for row in values[:100])))
                sheet.column_dimensions[get_column_letter(index)].width = width
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=root) as handle:
            temp_path = Path(handle.name)
        try:
            wb.save(temp_path)
            temp_path.replace(xlsx_path)
        finally:
            temp_path.unlink(missing_ok=True)
    except ImportError:
        xlsx_path = Path("")

    files = [json_path, md_path, tasks_csv, providers_csv, iterations_csv]
    if xlsx_path and xlsx_path.is_file():
        files.append(xlsx_path)
    hashes = {
        path.name: sha256_file(path, allowed_roots=[root], max_bytes=None)
        for path in files
    }
    atomic_json(root / "MANIFEST.json", {"schema": "iot-ai.autonomous-run-report-manifest.v2", "files": hashes})
    return {
        "decision": "pass",
        "root": str(root),
        "files": [str(path) for path in files],
        "manifest": str(root / "MANIFEST.json"),
        "report_digest": report["digest"],
    }
