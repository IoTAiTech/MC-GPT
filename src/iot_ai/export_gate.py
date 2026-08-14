# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Required Notice: Copyright 2026 IoT-AI.Tech / Dr.-Ing. Babak Sorkhpour
# Author: Dr.-Ing. Babak Sorkhpour, with AI assistance
# Version: 6.7.0-beta.5 | Date: 2026-08-08
"""Fail-closed redaction and classification gate for exported text artifacts."""
from __future__ import annotations
import hashlib
import io
import re
import zipfile
from pathlib import Path
from .privacy import sanitize
from .util import confined_text_read, confined_text_write, open_secure, resolve_within_allowed_roots, trusted_operator_roots

PRIVATE_IP = re.compile(r"\b(?:10|127)\.\d{1,3}\.\d{1,3}\.\d{1,3}\b|\b192\.168\.\d{1,3}\.\d{1,3}\b|\b172\.(?:1[6-9]|2\d|3[01])\.\d{1,3}\.\d{1,3}\b")
PRIVATE_PATH = re.compile(r"(?:/(?:home|root)/[A-Za-z0-9._-]+(?:/[^\s\"']+)?|[A-Za-z]:\\Users\\[^\r\n\"']+)")
SECRET_RESIDUAL = re.compile(r"(?i)(begin\s+(?:rsa\s+|ec\s+|openssh\s+)?private\s+key|akia[0-9a-z]{16}|ghp_[A-Za-z0-9]{20,}|github_pat_[A-Za-z0-9_]{20,}|xox[baprs]-)")

def redact_text(text: str) -> dict:
    findings: list[str] = []
    privacy = sanitize(text, mode="strict")
    out = privacy.text
    findings.extend(list(privacy.findings or ()))
    if PRIVATE_IP.search(out):
        findings.append("private_ip")
        out = PRIVATE_IP.sub("[PRIVATE_IP]", out)
    if PRIVATE_PATH.search(out):
        findings.append("private_path")
        out = PRIVATE_PATH.sub("[PRIVATE_PATH]", out)
    return {
        "text": out,
        "findings": sorted(set(findings)),
        "sha256": hashlib.sha256(out.encode("utf-8")).hexdigest(),
    }

def _findings_for_text(text: str) -> dict:
    result = redact_text(text)
    findings = list(result["findings"])
    if SECRET_RESIDUAL.search(text) or SECRET_RESIDUAL.search(result["text"]):
        findings.append("secret_residual")
    return {
        "decision": "pass" if not findings else "redact-required",
        "findings": sorted(set(findings)),
        "redacted_sha256": result["sha256"],
        "redacted_text": result["text"],
    }

def _read_confined_bytes(path: Path, roots: list[Path]) -> bytes:
    handle = open_secure(path, roots)
    try:
        return handle.read()
    finally:
        handle.close()

def _inspect_xlsx_bytes(data: bytes) -> dict:
    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        cells: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is not None:
                        cells.append(str(value))
    finally:
        workbook.close()
    return {**_findings_for_text("\n".join(cells)), "kind": "xlsx"}

def _inspect_zip_bytes(data: bytes) -> dict:
    findings: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                name = info.filename.lower()
                payload = archive.read(info)
                if name.endswith(".xlsx"):
                    inner = _inspect_xlsx_bytes(payload)
                    findings.extend(inner.get("findings") or [])
                    continue
                try:
                    text = payload.decode("utf-8")
                except UnicodeDecodeError:
                    findings.append("unknown_binary")
                    continue
                findings.extend(_findings_for_text(text)["findings"])
    except zipfile.BadZipFile:
        return {"decision": "block", "kind": "binary", "findings": ["unknown_binary"], "redacted_text": "", "redacted_sha256": hashlib.sha256(b"").hexdigest()}
    unique = sorted(set(findings))
    return {
        "decision": "block" if unique else "pass",
        "kind": "zip",
        "findings": unique,
        "redacted_text": "",
        "redacted_sha256": hashlib.sha256(b"").hexdigest(),
    }

def inspect_export_file(path: Path, *, allowed_roots: list[Path] | None = None) -> dict:
    roots = allowed_roots or list(trusted_operator_roots())
    safe = resolve_within_allowed_roots(path, roots, must_exist=True)
    suffix = Path(safe).suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        inspected = _inspect_zip_bytes(_read_confined_bytes(safe, roots))
        inspected["kind"] = "xlsx"
        return {"path": str(safe), **inspected}
    if suffix == ".zip":
        inspected = _inspect_zip_bytes(_read_confined_bytes(safe, roots))
        return {"path": str(safe), **inspected}
    try:
        text = confined_text_read(safe, roots)
    except UnicodeDecodeError:
        return {"decision": "block", "kind": "binary", "path": str(safe), "findings": ["unknown_binary"]}
    inspected = _findings_for_text(text)
    return {"path": str(safe), "kind": "text", **inspected}

def assert_export_safe(path: Path, *, public: bool = True, allowed_roots: list[Path] | None = None) -> dict:
    result = inspect_export_file(path, allowed_roots=allowed_roots)
    if public and result["decision"] != "pass":
        return {**result, "decision": "block"}
    return result

def rewrite_public_export(path: Path, *, allowed_roots: list[Path] | None = None) -> dict:
    roots = allowed_roots or list(trusted_operator_roots())
    safe = resolve_within_allowed_roots(path, roots, must_exist=True)
    result = inspect_export_file(safe, allowed_roots=roots)
    findings = set(result.get("findings") or [])
    kind = str(result.get("kind") or "text")
    if kind == "binary" or "unknown_binary" in findings or "secret_residual" in findings:
        return {k: v for k, v in {**result, "decision": "block"}.items() if k != "redacted_text"}
    if kind in {"xlsx", "zip"}:
        decision = "pass" if result.get("decision") == "pass" and not findings else "block"
        return {k: v for k, v in {**result, "decision": decision}.items() if k != "redacted_text"}
    if kind == "text" and result.get("decision") == "redact-required":
        confined_text_write(safe, str(result["redacted_text"]), roots)
        return {k: v for k, v in {**result, "decision": "pass"}.items() if k != "redacted_text"}
    return {k: v for k, v in result.items() if k != "redacted_text"}
