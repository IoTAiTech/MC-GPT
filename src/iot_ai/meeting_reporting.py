# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Required Notice: Copyright 2026 IoT-AI.Tech / Dr.-Ing. Babak Sorkhpour
# Author: Dr.-Ing. Babak Sorkhpour, with AI assistance
# Version: 6.7.0-beta.5 | Date: 2026-08-08
"""Federated, read-only Meeting reporting with fail-closed public export.

The Suite control database remains the sole canonical writer. Historical
meeting SQLite stores are optional operator-selected read-only inputs. Public
reports require explicit D0 allowlisting; restricted/private reports preserve
full evidence but are never public-release assets.
"""
from __future__ import annotations
import csv, hashlib, json, os, re, sqlite3, tempfile, zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence
from .export_gate import redact_text, rewrite_public_export
from .meeting import project_meeting_view, show
from .paths import data_root, db_path
from .util import assert_secure_regular_file, sha256_file, trusted_operator_roots, utc_now
from .workspace import connect_read, rows

ANSI_ESCAPE=re.compile(r"(?:\x1B[@-_][0-?]*[ -/]*[@-~]|\x1B\][^\x07]*(?:\x07|\x1B\\))")
DEFAULT_STALE_SECONDS=24*60*60
LEGACY_REQUIRED_TABLES={"meetings","contributions"}


def _clean_text(value: Any, limit: int|None=None)->str:
    text=ANSI_ESCAPE.sub("",str(value or "")).replace("\x00","")
    text="\n".join(line.rstrip() for line in text.splitlines()).strip()
    if limit is not None and len(text)>limit: return text[:max(0,limit-1)].rstrip()+"…"
    return text

def _parse_time(value: str|None)->datetime|None:
    if not value:return None
    try: parsed=datetime.fromisoformat(str(value).replace("Z","+00:00"))
    except ValueError:return None
    if parsed.tzinfo is None: parsed=parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)

def _stale(value: str|None,seconds:int)->bool:
    parsed=_parse_time(value)
    return bool(parsed and (datetime.now(timezone.utc)-parsed).total_seconds()>seconds)

def _meeting_id(item:dict[str,Any])->str:
    return str((item.get("meeting") or item).get("id") or item.get("meeting_id") or "")

def _privacy_class(item:dict[str,Any])->str:
    return str((item.get("meeting") or item).get("privacy_class") or item.get("privacy_class") or "D1")


def _canonical_ids(user_home:Path,**filters:Any)->list[str]:
    conn=connect_read(user_home)
    if conn is None:return []
    clauses=[];params=[]
    mapping={"from_time":("m.created_at>=?",filters.get("from_time")),"to_time":("m.created_at<=?",filters.get("to_time")),"status":("m.status=?",filters.get("status")),"task_id":("m.task_id=?",filters.get("task_id")),"decision":("m.final_decision=?",filters.get("decision"))}
    for _,(clause,value) in mapping.items():
        if value is not None: clauses.append(clause);params.append(value)
    provider=filters.get("provider")
    if provider:
        clauses.append("EXISTS(SELECT 1 FROM meeting_contributions c WHERE c.meeting_id=m.id AND (c.seat=? OR c.seat LIKE ?))");params.extend([provider,provider+"@%"])
    agent=filters.get("agent")
    if agent:
        clauses.append("EXISTS(SELECT 1 FROM meeting_contributions c WHERE c.meeting_id=m.id AND c.seat=?)");params.append(agent if str(agent).startswith("agent:") else "agent:"+str(agent))
    where=" WHERE "+" AND ".join(clauses) if clauses else ""
    result=[r["id"] for r in rows(conn,"SELECT m.id FROM meetings m"+where+" ORDER BY m.created_at,m.id",params)]
    conn.close();return result


def _canonical_collect(user_home:Path,view:str,stale_after_seconds:int,**filters:Any)->list[dict[str,Any]]:
    result=[]
    for meeting_id in _canonical_ids(user_home,**filters):
        item=project_meeting_view(show(user_home,meeting_id),view)
        meeting=item.get("meeting") or item
        source_status=str(meeting.get("status") or item.get("status") or "unknown")
        updated=meeting.get("updated_at") or item.get("updated_at")
        stale=source_status=="running" and _stale(updated,stale_after_seconds)
        if stale:
            if item.get("meeting"):item["meeting"]["status"]="stale"
            item["status"]="stale"
        participants=item.get("participants") or meeting.get("contributions") or []
        issues=[]
        if stale:issues.append("canonical-running-session-stale")
        approved=bool(item.get("founder_approval") or meeting.get("user_approved"))
        if approved and source_status not in {"approved","completed","closed","awaiting-user-decision"}:issues.append("approval-status-conflict")
        missing=sum(1 for p in participants if not p.get("model_served"))
        if missing:issues.append(f"model-served-unverified:{missing}/{len(participants)}")
        item.update({"privacy_class":str(meeting.get("privacy_class") or item.get("privacy_class") or "D1"),"source":{"kind":"canonical","label":"suite-control-database","source_status":source_status,"stale_reclassified":stale},"lifecycle_issues":issues,"model_telemetry_complete":missing==0})
        result.append(item)
    return result


def _legacy_connection(user_home:Path,path:Path)->tuple[sqlite3.Connection,Path]:
    roots=trusted_operator_roots(user_home)
    resolved=assert_secure_regular_file(path,roots,max_bytes=256*1024*1024)
    conn=sqlite3.connect(f"file:{resolved}?mode=ro",uri=True);conn.row_factory=sqlite3.Row;conn.execute("PRAGMA query_only=ON")
    tables={r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    if not LEGACY_REQUIRED_TABLES.issubset(tables):conn.close();raise ValueError(f"unsupported legacy meeting schema: {path}")
    check=conn.execute("PRAGMA integrity_check").fetchone()
    if not check or str(check[0]).lower()!="ok":conn.close();raise ValueError(f"legacy meeting database failed integrity_check: {path}")
    return conn,resolved

def _legacy_decision(row:dict[str,Any])->str:
    text=_clean_text(row.get("final_text"),500)
    match=re.search(r"(?im)^\s*(?:DECISION|RESULT)\s*:\s*([A-Za-z_-]+)",text)
    if match:return match.group(1).lower().replace("_","-")
    return {"blocked":"block","needs-review":"needs-work","approved":"approve","completed":"pass","running":"pending"}.get(str(row.get("status") or "unknown"),str(row.get("status") or "unknown"))

def _legacy_participants(conn:sqlite3.Connection,meeting_id:str)->list[dict[str,Any]]:
    grouped={}
    for raw in conn.execute("SELECT * FROM contributions WHERE meeting_id=? ORDER BY round_no,created_at,contribution_id",(meeting_id,)):
        row=dict(raw);seat=str(row.get("seat_id") or row.get("provider") or "unknown")
        p=grouped.setdefault(seat,{"seat":seat,"provider":row.get("provider") or seat.split("@",1)[0],"status":"not-observed","decision":None,"model_requested":None,"model_served":None,"quality_score":None,"failure_class":None,"opinion_summary":"","critique_count":0,"final_review_summary":"","contribution_count":0,"substantive":False,"legacy_model_telemetry_unverified":False})
        p["contribution_count"]+=1;p["status"]=row.get("status") or p["status"];p["decision"]=row.get("decision") or p["decision"]
        p["model_requested"]=row.get("model_requested") or p["model_requested"];p["model_served"]=row.get("model_served") or p["model_served"]
        content=_clean_text(row.get("content"),1200);stage=str(row.get("stage") or "")
        if stage=="opinion" and not p["opinion_summary"]:p["opinion_summary"]=content
        elif "critique" in stage:p["critique_count"]+=1
        elif stage in {"final-review","revision"}:p["final_review_summary"]=content
        status=str(row.get("status") or "").lower()
        if status in {"pass","completed"} and len(content)>=40 and row.get("model_served"):p["substantive"]=True
    for p in grouped.values():p["legacy_model_telemetry_unverified"]=not bool(p.get("model_served"))
    return [grouped[k] for k in sorted(grouped)]

def _legacy_collect(user_home:Path,path:Path,label:str,view:str,stale_after_seconds:int,**filters:Any)->list[dict[str,Any]]:
    conn,resolved=_legacy_connection(user_home,path)
    result=[]
    try:
        clauses=[];params=[]
        for key,column,op in (("from_time","created_at",">="),("to_time","created_at","<="),("status","status","="),("task_id","task_id","=")):
            value=filters.get(key)
            if value is not None:clauses.append(f"{column}{op}?");params.append(value)
        where=" WHERE "+" AND ".join(clauses) if clauses else ""
        for raw in conn.execute("SELECT * FROM meetings"+where+" ORDER BY created_at,meeting_id",params):
            row=dict(raw);meeting_id=str(row["meeting_id"]);participants=_legacy_participants(conn,meeting_id)
            provider=filters.get("provider");agent=filters.get("agent");decision=_legacy_decision(row)
            if provider and not any(str(p.get("provider"))==str(provider) for p in participants):continue
            if agent and not any(str(p.get("seat"))==(str(agent) if str(agent).startswith("agent:") else "agent:"+str(agent)) for p in participants):continue
            if filters.get("decision") and decision!=filters["decision"]:continue
            source_status=str(row.get("status") or "unknown");stale=source_status=="running" and _stale(row.get("updated_at"),stale_after_seconds)
            status="stale" if stale else source_status;approved=bool(row.get("user_approved"));issues=[]
            if stale:issues.append("legacy-running-session-stale")
            if approved and source_status not in {"approved","completed","closed","awaiting-user-decision"}:issues.append("approval-status-conflict")
            missing=sum(1 for p in participants if not p.get("model_served"))
            if missing:issues.append(f"legacy-model-served-unverified:{missing}/{len(participants)}")
            substantive=sum(1 for p in participants if p.get("substantive"))
            seats=json.loads(row.get("seats_json") or "[]")
            requested=[s.get("seat_id") if isinstance(s,dict) else s for s in seats]
            item={"view":view,"meeting_id":meeting_id,"task_id":row.get("task_id"),"created_at":row.get("created_at"),"updated_at":row.get("updated_at"),"privacy_class":row.get("privacy_class") or "D1","mode":row.get("mode"),"topic":_clean_text(row.get("topic")),"topic_preview":_clean_text(row.get("topic"),400),"status":status,"source_status":source_status,"final_decision":decision,"founder_approval":approved,"requested_seats":requested,"substantive_seats":substantive,"synthesizer_seat":row.get("synthesizer_seat"),"synthesis_summary":_clean_text(row.get("final_text"),None if view=="full" else 1200),"participants":participants,"blockers":issues if decision in {"block","needs-work","pending"} else [],"lifecycle_issues":issues,"model_telemetry_complete":missing==0,"source":{"kind":"legacy-read-only","label":label,"source_id":"legacy-"+hashlib.sha256(str(resolved).encode()).hexdigest()[:12],"source_status":source_status,"stale_reclassified":stale}}
            result.append(item)
    finally:conn.close()
    return result


def _public_filter(meetings:list[dict[str,Any]],allowlist:set[str])->tuple[list[dict[str,Any]],list[dict[str,str]],list[str]]:
    included=[];omitted=[];findings=[]
    for item in meetings:
        mid=_meeting_id(item)
        if mid not in allowlist:omitted.append({"meeting_id":mid,"reason":"not-explicitly-allowlisted"});continue
        if _privacy_class(item)!="D0":omitted.append({"meeting_id":mid,"reason":"privacy-class-not-D0"});continue
        meeting=item.get("meeting") or item
        topic=redact_text(_clean_text(meeting.get("topic") or item.get("topic") or item.get("topic_preview"),400));findings.extend(topic["findings"])
        participants=item.get("participants") or meeting.get("contributions") or []
        safe_participants=[]
        for p in participants:
            safe_participants.append({"seat":p.get("seat") or p.get("seat_id"),"provider":p.get("provider") or str(p.get("seat") or "").split("@",1)[0],"status":p.get("status"),"decision":p.get("decision"),"model_served":p.get("model_served") or "unverified","legacy_model_telemetry_unverified":not bool(p.get("model_served"))})
        included.append({"view":"brief","meeting_id":mid,"task_id":meeting.get("task_id") or item.get("task_id"),"created_at":meeting.get("created_at") or item.get("created_at"),"updated_at":meeting.get("updated_at") or item.get("updated_at"),"privacy_class":"D0","topic_preview":topic["text"],"status":meeting.get("status") or item.get("status"),"final_decision":meeting.get("final_decision") or item.get("final_decision"),"founder_approval":bool(item.get("founder_approval") or meeting.get("user_approved")),"requested_seats":meeting.get("requested_seats") or item.get("requested_seats"),"substantive_seats":meeting.get("substantive_seats") or item.get("substantive_seats"),"participants":safe_participants,"blockers":item.get("blockers") or [],"lifecycle_issues":item.get("lifecycle_issues") or [],"model_telemetry_complete":item.get("model_telemetry_complete"),"source":{"kind":(item.get("source") or {}).get("kind"),"label":(item.get("source") or {}).get("label"),"source_status":(item.get("source") or {}).get("source_status"),"stale_reclassified":(item.get("source") or {}).get("stale_reclassified")}})
    return included,omitted,sorted(set(findings))


def _source_manifest(user_home:Path,legacy:list[Path],public:bool,include_current:bool)->list[dict[str,Any]]:
    sources=[];canonical=db_path(user_home)
    if include_current and canonical.is_file():sources.append({"label":"suite-control-database","kind":"canonical","sha256":sha256_file(canonical,allowed_roots=[user_home],max_bytes=None),**({} if public else {"path":str(canonical)})})
    for i,path in enumerate(legacy,1):
        resolved=assert_secure_regular_file(path,trusted_operator_roots(user_home),max_bytes=256*1024*1024)
        sources.append({"label":f"legacy-{i}","kind":"legacy-read-only","source_id":"legacy-"+hashlib.sha256(str(resolved).encode()).hexdigest()[:12],"sha256":sha256_file(resolved,allowed_roots=trusted_operator_roots(user_home),max_bytes=None),**({} if public else {"path":str(resolved)})})
    return sources

def _model_summary(meetings:list[dict[str,Any]])->list[dict[str,Any]]:
    counts=defaultdict(Counter)
    for item in meetings:
        participants=item.get("participants") or (item.get("meeting") or {}).get("contributions") or []
        for p in participants:
            provider=str(p.get("provider") or p.get("seat") or "unknown").split("@",1)[0];model=str(p.get("model_served") or "unverified");c=counts[(provider,model)];c["attempted"]+=1
            if str(p.get("status")) in {"pass","completed"}:c["completed"]+=1
            else:c["failed_or_other"]+=1
            if p.get("substantive") or p.get("opinion_summary"):c["substantive"]+=1
            if not p.get("model_served"):c["model_unverified"]+=1
    return [{"provider":p,"model_served":m,**dict(c)} for (p,m),c in sorted(counts.items())]
def _status_summary(meetings:list[dict[str,Any]])->dict[str,Any]:
    statuses=Counter(str((i.get("meeting") or i).get("status") or i.get("status") or "unknown") for i in meetings);decisions=Counter(str((i.get("meeting") or i).get("final_decision") or i.get("final_decision") or "unknown") for i in meetings)
    return {"statuses":dict(sorted(statuses.items())),"decisions":dict(sorted(decisions.items())),"privacy_classes":dict(sorted(Counter(_privacy_class(i) for i in meetings).items())),"sources":dict(sorted(Counter(str((i.get("source") or {}).get("label") or "canonical") for i in meetings).items())),"stale_reclassified":sum(bool((i.get("source") or {}).get("stale_reclassified")) for i in meetings),"founder_approved":sum(bool(i.get("founder_approval") or (i.get("meeting") or {}).get("user_approved")) for i in meetings)}


def collect(user_home:Path,*,view:str="brief",legacy_dbs:Sequence[Path]|None=None,legacy_stores:Sequence[Path|str]|None=None,public:bool=False,classification:str|None=None,public_allowlist:Sequence[str]|None=None,stale_after_seconds:int=DEFAULT_STALE_SECONDS,stale_after_hours:int|None=None,include_current:bool=True,**filters:Any)->dict[str,Any]:
    if classification:
        c=classification.lower()
        if c not in {"public","private","restricted"}:raise ValueError("classification must be public, private or restricted")
        public=c=="public"
    if stale_after_hours is not None:stale_after_seconds=int(stale_after_hours)*3600
    mode="brief" if view in {"brief","simple"} else "full"
    if public and mode!="brief":raise PermissionError("public meeting reports are brief-only")
    legacy=[Path(v) for v in [*(legacy_dbs or []),*(legacy_stores or [])]]
    meetings=_canonical_collect(user_home,mode,stale_after_seconds,**filters) if include_current else []
    for i,path in enumerate(legacy,1):meetings.extend(_legacy_collect(user_home,path,f"legacy-{i}",mode,stale_after_seconds,**filters))
    meetings.sort(key=lambda i:(str((i.get("meeting") or i).get("created_at") or i.get("created_at") or ""),_meeting_id(i)))
    omitted=[];findings=[]
    if public:
        allow={str(v) for v in (public_allowlist or []) if str(v)}
        if not allow:raise PermissionError("public report requires at least one explicit public meeting allowlist entry")
        meetings,omitted,findings=_public_filter(meetings,allow)
        if not meetings:raise PermissionError("public report contains no explicitly approved D0 meetings after filtering")
    payload={"schema":"iot-ai.cross-meeting-report.v2","generated_at":utc_now(),"view":mode,"classification":"PUBLIC-SANITIZED" if public else ("RESTRICTED-FORENSIC" if (classification or "").lower()=="restricted" else "PRIVATE-CONFIDENTIAL"),"public_export":public,"filters":{k:v for k,v in filters.items() if v is not None},"source_manifest":_source_manifest(user_home,legacy,public,include_current),"meeting_count":len(meetings),"omitted_count":len(omitted),"omitted":omitted,"public_redaction_findings":findings,"summary":_status_summary(meetings),"model_participation":_model_summary(meetings),"lifecycle_issues":[{"meeting_id":_meeting_id(i),"source":(i.get("source") or {}).get("label"),"issue":issue} for i in meetings for issue in (i.get("lifecycle_issues") or [])],"claims":{"legacy_model_identity_inferred":False,"missing_model_served_counts_as_qualified":False,"global_compliance_claim_allowed":False},"meetings":meetings}
    canonical=json.dumps(payload,ensure_ascii=False,sort_keys=True,separators=(",",":")).encode();payload["report_payload_sha256"]=hashlib.sha256(canonical).hexdigest();return payload


def _flat(payload:dict[str,Any])->list[dict[str,Any]]:
    output=[]
    for item in payload["meetings"]:
        m=item.get("meeting") or item;parts=item.get("participants") or m.get("contributions") or []
        output.append({"meeting_id":m.get("id") or item.get("meeting_id"),"source":(item.get("source") or {}).get("label"),"task_id":m.get("task_id") or item.get("task_id"),"created_at":m.get("created_at") or item.get("created_at"),"updated_at":m.get("updated_at") or item.get("updated_at"),"privacy_class":m.get("privacy_class") or item.get("privacy_class"),"topic":m.get("topic") or item.get("topic") or item.get("topic_preview"),"status":m.get("status") or item.get("status"),"source_status":(item.get("source") or {}).get("source_status"),"stale_reclassified":bool((item.get("source") or {}).get("stale_reclassified")),"final_decision":m.get("final_decision") or item.get("final_decision"),"founder_approval":bool(item.get("founder_approval") or m.get("user_approved")),"requested_seats":json.dumps(m.get("requested_seats") or item.get("requested_seats") or [],ensure_ascii=False),"substantive_seats":m.get("substantive_seats") or item.get("substantive_seats"),"model_telemetry_complete":item.get("model_telemetry_complete"),"lifecycle_issues":json.dumps(item.get("lifecycle_issues") or [],ensure_ascii=False),"participant_models":json.dumps([{"seat":p.get("seat") or p.get("seat_id"),"model_requested":p.get("model_requested"),"model_served":p.get("model_served"),"status":p.get("status"),"decision":p.get("decision")} for p in parts],ensure_ascii=False),"synthesis_summary":item.get("synthesis_summary") or m.get("synthesis"),"blockers":json.dumps(item.get("blockers") or [],ensure_ascii=False)})
    return output

def _participant_rows(payload:dict[str,Any])->list[dict[str,Any]]:
    out=[]
    for item in payload["meetings"]:
        mid=_meeting_id(item);parts=item.get("participants") or (item.get("meeting") or {}).get("contributions") or []
        for p in parts:out.append({"meeting_id":mid,"source":(item.get("source") or {}).get("label"),"seat":p.get("seat") or p.get("seat_id"),"provider":p.get("provider"),"status":p.get("status"),"decision":p.get("decision"),"model_requested":p.get("model_requested"),"model_served":p.get("model_served"),"substantive":p.get("substantive"),"legacy_model_telemetry_unverified":p.get("legacy_model_telemetry_unverified"),"opinion_summary":p.get("opinion_summary"),"critique_count":p.get("critique_count"),"final_review_summary":p.get("final_review_summary")})
    return out

def _decision_rows(payload:dict[str,Any])->list[dict[str,Any]]:
    return [{"meeting_id":_meeting_id(i),"status":(i.get("meeting") or i).get("status") or i.get("status"),"decision":(i.get("meeting") or i).get("final_decision") or i.get("final_decision"),"founder_approval":bool(i.get("founder_approval") or (i.get("meeting") or {}).get("user_approved")),"blockers":json.dumps(i.get("blockers") or [],ensure_ascii=False)} for i in payload["meetings"]]
def _write_csv(path:Path,data:list[dict[str,Any]],fallback:list[str])->None:
    fields=list(data[0]) if data else fallback
    with path.open("w",encoding="utf-8",newline="") as h:w=csv.DictWriter(h,fieldnames=fields,extrasaction="ignore");w.writeheader();w.writerows(data)
def _write_markdown(path:Path,payload:dict[str,Any])->None:
    lines=["# IOT-AI Meeting Report","",f"Classification: `{payload['classification']}`",f"Generated: {payload['generated_at']}",f"Meetings: {payload['meeting_count']}",""]
    for item in payload["meetings"]:
        m=item.get("meeting") or item;lines.extend([f"## {_meeting_id(item)}","",f"- Source: `{(item.get('source') or {}).get('label')}`",f"- Status: `{m.get('status') or item.get('status')}`",f"- Decision: `{m.get('final_decision') or item.get('final_decision')}`",f"- Privacy: `{_privacy_class(item)}`",f"- Topic: {m.get('topic') or item.get('topic') or item.get('topic_preview') or ''}",f"- Lifecycle issues: `{', '.join(item.get('lifecycle_issues') or []) or 'none'}`",""])
        for p in item.get("participants") or []:lines.append(f"- **{p.get('seat')}** · {p.get('status')} · {p.get('model_served') or 'unverified'} · {p.get('decision') or ''}")
        if not payload["public_export"]:lines.extend(["",item.get("synthesis_summary") or m.get("synthesis") or "",""])
    path.write_text("\n".join(lines).rstrip()+"\n",encoding="utf-8")
def _style_sheet(ws,widths:dict[str,int]|None=None)->None:
    from openpyxl.styles import Font,PatternFill,Alignment
    for cell in ws[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="17365D");cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for c in row:c.alignment=Alignment(vertical="top",wrap_text=True)
    for col,w in (widths or {}).items():ws.column_dimensions[col].width=w
def _write_xlsx(path:Path,payload:dict[str,Any])->None:
    from openpyxl import Workbook
    wb=Workbook();ws=wb.active;ws.title="Meetings";data=_flat(payload);heads=list(data[0]) if data else ["meeting_id","status"];ws.append(heads)
    for r in data:ws.append([r.get(h) for h in heads])
    _style_sheet(ws,{"A":34,"B":16,"C":34,"D":24,"E":24,"F":12,"G":48,"H":18,"I":18,"J":12,"K":18,"L":14,"M":14,"N":56,"O":56,"P":28})
    part=wb.create_sheet("Participants");pdata=_participant_rows(payload);ph=list(pdata[0]) if pdata else ["meeting_id","seat","status","model_served"];part.append(ph)
    for r in pdata:part.append([r.get(h) for h in ph])
    _style_sheet(part,{"A":34,"B":16,"C":20,"D":16,"E":16,"F":16,"G":28,"H":28,"I":14,"J":18,"K":56,"L":14,"M":56})
    models=wb.create_sheet("Model Participation");md=payload["model_participation"];mh=list(md[0]) if md else ["provider","model_served","attempted"];models.append(mh)
    for r in md:models.append([r.get(h,0) for h in mh])
    _style_sheet(models,{"A":18,"B":40,"C":14,"D":14,"E":18,"F":14,"G":18})
    summary=wb.create_sheet("Summary");summary.append(["Metric","Value"]);summary.append(["Classification",payload["classification"]]);summary.append(["Generated",payload["generated_at"]]);summary.append(["Meetings included",payload["meeting_count"]]);summary.append(["Meetings omitted",payload["omitted_count"]]);summary.append(["Payload SHA-256",payload["report_payload_sha256"]])
    for group,values in payload["summary"].items():summary.append([group,json.dumps(values,ensure_ascii=False,sort_keys=True) if isinstance(values,dict) else values])
    _style_sheet(summary,{"A":28,"B":90})
    fd,tmp=tempfile.mkstemp(prefix=".meeting-report-",suffix=".xlsx",dir=str(path.parent));os.close(fd)
    try:wb.save(tmp);os.replace(tmp,path)
    finally:Path(tmp).unlink(missing_ok=True)

def managed_report_output(user_home:Path,filename:str)->Path:
    if not filename or Path(filename).name!=filename or any(ch not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-" for ch in filename):raise ValueError("invalid report filename")
    root=(data_root(user_home)/"meeting-reports").resolve();root.mkdir(parents=True,exist_ok=True);output=(root/filename).resolve()
    if output.parent!=root:raise PermissionError("report output escaped managed root")
    return output

def _manifest(output:Path,payload:dict[str,Any],gate:dict[str,Any]|None)->dict[str,Any]:
    return {"schema":"iot-ai.meeting-report-file.v2","classification":payload["classification"],"generated_at":payload["generated_at"],"report_payload_sha256":payload["report_payload_sha256"],"meeting_count":payload["meeting_count"],"file":output.name,"public_export":gate,"source_manifest":payload["source_manifest"]}

def write_report(user_home:Path,output:Path,*,output_format:str,view:str="brief",**kwargs:Any)->dict[str,Any]:
    fmt=output_format.lower()
    if fmt in {"bundle","all"}:return write_report_bundle(user_home,output,view=view,**kwargs)
    payload=collect(user_home,view=view,**kwargs);output.parent.mkdir(parents=True,exist_ok=True)
    if fmt=="json":output.write_text(json.dumps(payload,ensure_ascii=False,indent=2,sort_keys=True)+"\n",encoding="utf-8")
    elif fmt=="csv":_write_csv(output,_flat(payload),["meeting_id","status","final_decision"])
    elif fmt in {"md","markdown"}:_write_markdown(output,payload)
    elif fmt=="xlsx":_write_xlsx(output,payload)
    else:raise ValueError("format must be json, csv, markdown, xlsx or bundle")
    public=bool(payload["public_export"]);gate=rewrite_public_export(output) if public else None
    if public and (gate or {}).get("decision")!="pass":output.unlink(missing_ok=True);raise PermissionError(f"public report export blocked: {(gate or {}).get('findings')}")
    manifest=output.with_name(output.name+(".public-manifest.json" if public else ".manifest.json"));manifest.write_text(json.dumps(_manifest(output,payload,gate),ensure_ascii=False,indent=2,sort_keys=True)+"\n",encoding="utf-8")
    digest=sha256_file(output,allowed_roots=[user_home,output.parent.resolve()],max_bytes=None);side=output.with_name(output.name+".sha256");side.write_text(f"{digest}  {output.name}\n",encoding="utf-8")
    return {"decision":"pass","output":str(output),"format":fmt,"view":payload["view"],"classification":payload["classification"],"meeting_count":payload["meeting_count"],"omitted_count":payload["omitted_count"],"sha256":digest,"sha256_sidecar":str(side),"manifest":str(manifest),"manifest_sha256":sha256_file(manifest,allowed_roots=[user_home,manifest.parent.resolve()],max_bytes=None),"public_export":gate}

def write_report_bundle(user_home:Path,output_dir:Path,*,view:str="brief",**kwargs:Any)->dict[str,Any]:
    output_dir=Path(output_dir)
    if output_dir.suffix.lower()==".zip":
        output_dir.parent.mkdir(parents=True,exist_ok=True);stage=Path(tempfile.mkdtemp(prefix=".iot-ai-meeting-bundle-",dir=str(output_dir.parent)))
        try:
            result=write_report_bundle(user_home,stage,view=view,**kwargs);tmp=output_dir.with_name("."+output_dir.name+".tmp")
            with zipfile.ZipFile(tmp,"w",zipfile.ZIP_DEFLATED,compresslevel=9) as archive:
                for p in sorted(stage.iterdir()):
                    if p.is_file():
                        info=zipfile.ZipInfo(p.name,date_time=(2026,8,8,0,0,0));info.compress_type=zipfile.ZIP_DEFLATED;info.external_attr=0o644<<16;archive.writestr(info,p.read_bytes())
            os.replace(tmp,output_dir);digest=sha256_file(output_dir,allowed_roots=[user_home,output_dir.parent.resolve()],max_bytes=None);side=output_dir.with_name(output_dir.name+".sha256");side.write_text(f"{digest}  {output_dir.name}\n",encoding="utf-8");members=sorted(p.name for p in stage.iterdir() if p.is_file())
            return {"decision":"pass","output":str(output_dir),"archive":True,"classification":result["classification"],"meeting_count":result["meeting_count"],"omitted_count":result["omitted_count"],"files":len(members),"archive_members":members,"internal_manifest":"MANIFEST.json","internal_checksums":"SHA256SUMS.txt","sha256":digest,"sha256_sidecar":str(side)}
        finally:
            for p in sorted(stage.rglob("*"),reverse=True):
                if p.is_file():p.unlink(missing_ok=True)
                elif p.is_dir():p.rmdir()
            stage.rmdir()
    output_dir.mkdir(parents=True,exist_ok=True);payload=collect(user_home,view=view,**kwargs);public=bool(payload["public_export"])
    files={"MEETINGS_INDEX.json":"json","MEETINGS_SUMMARY.csv":"csv","MEETINGS_REPORT.md":"markdown","MEETINGS_REPORT.xlsx":"xlsx"};results=[]
    for name,fmt in files.items():
        p=output_dir/name
        if fmt=="json":p.write_text(json.dumps(payload,ensure_ascii=False,indent=2,sort_keys=True)+"\n",encoding="utf-8")
        elif fmt=="csv":_write_csv(p,_flat(payload),["meeting_id","status","final_decision"])
        elif fmt=="markdown":_write_markdown(p,payload)
        else:_write_xlsx(p,payload)
        gate=rewrite_public_export(p) if public else None
        if public and (gate or {}).get("decision")!="pass":raise PermissionError(f"public report bundle blocked for {name}")
        results.append({"path":name,"sha256":sha256_file(p,allowed_roots=[user_home,output_dir.resolve()],max_bytes=None),"format":fmt})
    extras={"MODEL_PARTICIPATION.csv":payload["model_participation"],"DECISIONS_AND_DISSENTS.csv":_decision_rows(payload),"LIFECYCLE_ISSUES.csv":payload["lifecycle_issues"]}
    for name,data in extras.items():
        p=output_dir/name;fallback=["provider","model_served","attempted"] if name.startswith("MODEL") else ["meeting_id","status","decision","blockers"] if name.startswith("DECISIONS") else ["meeting_id","source","issue"]
        _write_csv(p,data,fallback);results.append({"path":name,"sha256":sha256_file(p,allowed_roots=[user_home,output_dir.resolve()],max_bytes=None),"format":"csv"})
    provenance=output_dir/"PROVENANCE.json";provenance.write_text(json.dumps({"schema":"iot-ai.meeting-report-provenance.v2","generated_at":payload["generated_at"],"classification":payload["classification"],"source_manifest":payload["source_manifest"],"claims":payload["claims"],"ansi_control_characters_removed":True,"legacy_stores_read_only":True},ensure_ascii=False,indent=2,sort_keys=True)+"\n",encoding="utf-8");results.append({"path":provenance.name,"sha256":sha256_file(provenance,allowed_roots=[user_home,output_dir.resolve()],max_bytes=None),"format":"json"})
    manifest={"schema":"iot-ai.meeting-report-bundle.v2","classification":payload["classification"],"generated_at":payload["generated_at"],"report_payload_sha256":payload["report_payload_sha256"],"meeting_count":payload["meeting_count"],"omitted_count":payload["omitted_count"],"included_meeting_ids":[_meeting_id(i) for i in payload["meetings"]],"omitted":payload["omitted"],"source_manifest":payload["source_manifest"],"files":results};text=json.dumps(manifest,ensure_ascii=False,indent=2,sort_keys=True)+"\n"
    for name in (["PUBLIC_REPORT_MANIFEST.json","MANIFEST.json"] if public else ["REPORT_MANIFEST.json","MANIFEST.json"]):(output_dir/name).write_text(text,encoding="utf-8")
    sums=output_dir/"SHA256SUMS.txt";sums.write_text("".join(f"{sha256_file(p,allowed_roots=[user_home,output_dir.resolve()],max_bytes=None)}  {p.name}\n" for p in sorted(output_dir.iterdir()) if p.is_file() and p.name!="SHA256SUMS.txt"),encoding="utf-8")
    return {"decision":"pass","output":str(output_dir),"classification":payload["classification"],"meeting_count":payload["meeting_count"],"omitted_count":payload["omitted_count"],"files":len([p for p in output_dir.iterdir() if p.is_file()]),"manifest":str(output_dir/"MANIFEST.json"),"manifest_sha256":sha256_file(output_dir/"MANIFEST.json",allowed_roots=[user_home,output_dir.resolve()],max_bytes=None),"checksums":str(sums),"checksums_sha256":sha256_file(sums,allowed_roots=[user_home,output_dir.resolve()],max_bytes=None)}
