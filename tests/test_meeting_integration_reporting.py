# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Required Notice: Copyright 2026 IoT-AI.Tech / Dr.-Ing. Babak Sorkhpour
# Author: Dr.-Ing. Babak Sorkhpour, with AI assistance
# Version: 6.7.0-beta.5 | Date: 2026-08-08
from __future__ import annotations
import hashlib, http.client, json, os, threading, unittest
from http.server import ThreadingHTTPServer
from pathlib import Path
from unittest.mock import patch
from openpyxl import load_workbook
from iot_ai.agent_seats import build_agent_envelope, validate_agent_reply
from iot_ai.installer import HOSTS, PUBLIC_SKILLS, install, verify
from iot_ai.meeting import start
from iot_ai.meeting_api import _handler
from iot_ai.meeting_integration import create_calendar_event, list_agent_seats, list_calendar_events, register_agent_seat
from iot_ai.meeting_reporting import collect, managed_report_output, write_report
from iot_ai.seat_selection import resolve_meeting_seats
from tests.common import IsolatedHomeTestCase

class MeetingIntegrationReportingTests(IsolatedHomeTestCase):
    def test_calendar_and_agent_registry(self):
        event=create_calendar_event(self.home,title="Review",topic="Review architecture",starts_at="2026-08-08T09:00:00+02:00",requested_seats="claude,agent:pmd/security",created_by="test",surface="pmd")
        self.assertEqual(event["starts_at"],"2026-08-08T07:00:00Z")
        self.assertEqual(len(list_calendar_events(self.home)),1)
        seat=register_agent_seat(self.home,surface="pmd",agent_id="security",display_name="Security",model_binding="local-model",endpoint_ref="http://127.0.0.1:9000/consult",capabilities=["review"],reachable=True)
        self.assertEqual(seat["seat"],"agent:pmd/security")
        self.assertEqual(list_agent_seats(self.home,surface="pmd")[0]["capabilities"],["review"])

    def test_agent_envelope_is_read_only_and_rejects_writes(self):
        envelope=build_agent_envelope("agent:pmd/security","Review","opinion","meeting-1","reviewer",30)
        text="Evidence-backed read-only assessment."
        reply={"status":"pass","text":text,"text_sha256":hashlib.sha256(text.encode()).hexdigest(),"model_served":"local-model","envelope_id":envelope["envelope_id"],"envelope_sha256":envelope["envelope_sha256"],"writes_performed":1}
        result=validate_agent_reply(envelope,reply)
        self.assertEqual(result["status"],"failed"); self.assertEqual(result["failure_class"],"policy_violation")

    def test_cross_meeting_all_formats_use_canonical_store(self):
        start(self.home,"First review",["claude"],quorum=1)
        start(self.home,"Second review",["codex"],quorum=1)
        self.assertEqual(collect(self.home,view="brief")["meeting_count"],2)
        written={}
        for fmt,suffix in (("json","json"),("csv","csv"),("markdown","md"),("xlsx","xlsx")):
            requested=self.home/f"report.{suffix}"
            result=write_report(self.home,requested,output_format=fmt,view="brief")
            stored=Path(result["output"])
            canonical=managed_report_output(self.home, f"report.{suffix}")
            self.assertEqual(result["meeting_count"],2)
            self.assertEqual(stored, canonical)
            self.assertTrue(stored.is_file(), msg=stored)
            self.assertNotEqual(stored, requested)
            written[fmt]=stored
        self.assertEqual(load_workbook(written["xlsx"])["Meetings"].max_row,3)
        with self.assertRaises(ValueError): managed_report_output(self.home,"../escape.json")

    @patch("iot_ai.seat_selection._ollama_cloud_seats", return_value=([], []))
    @patch("iot_ai.seat_selection._provider_routes")
    @patch("iot_ai.seat_selection.provider_candidates")
    def test_all_coders_selector_includes_every_exact_cloud_model(self, candidates, routes, _ollama):
        routes.return_value = [
            {"provider": "claude", "route_id": "claude-route", "installed": True, "cloud": True, "priority": 1},
            {"provider": "codex", "route_id": "codex-route", "installed": True, "cloud": True, "priority": 1},
        ]
        all_candidates = [
            {"provider": "claude", "route_id": "claude-route", "model": "model-a", "cloud": True, "live_ready": True},
            {"provider": "claude", "route_id": "claude-route", "model": "model-b", "cloud": True, "live_ready": False},
            {"provider": "codex", "route_id": "codex-route", "model": "auto", "cloud": True, "live_ready": False},
        ]
        candidates.side_effect = lambda *_args, **kwargs: [
            row for row in all_candidates if (row["live_ready"] or not kwargs.get("require_live"))
        ]
        plan = resolve_meeting_seats(self.home, "all-coders+ollama-clouds", allow_missing_ollama=True, max_seats=8)
        self.assertEqual(plan.decision, "pass")
        self.assertEqual(list(plan.resolved_seats), ["claude@model-a", "claude@model-b", "codex"])
        status = {row["seat"]: row for row in plan.candidate_status}
        self.assertTrue(status["claude@model-a"]["live_ready"])
        self.assertFalse(status["claude@model-b"]["live_ready"])

    @patch("iot_ai.seat_selection._coder_seats",return_value=([],[]))
    @patch("iot_ai.seat_selection._ollama_cloud_seats",return_value=([],[]))
    @patch("iot_ai.seat_selection._agent_candidates",return_value=([],[]))
    @patch("iot_ai.seat_selection._qualified_cloud_seats")
    def test_all_qualified_models_no_hidden_truncation(self,qualified,*_):
        rows=[{"seat":"claude@model-a","provider":"claude","live_ready":True},{"seat":"grok@model-b","provider":"grok","live_ready":True}]
        qualified.return_value=([r["seat"] for r in rows],rows)
        plan=resolve_meeting_seats(self.home,"all-qualified-cloud-models",max_seats=2)
        self.assertEqual(list(plan.resolved_seats),["claude@model-a","grok@model-b"])
        blocked=resolve_meeting_seats(self.home,"all-qualified-cloud-models",max_seats=1)
        self.assertEqual(blocked.decision,"block"); self.assertIn("SEAT_LIMIT_EXCEEDED",blocked.reason or "")

    def test_host_install_discovers_governed_commands(self):
        result=install(self.home,list(HOSTS)); self.assertEqual(result["decision"],"pass")
        self.assertTrue({"iot-ai-meeting","iot-ai-tasks","iot-ai-multi-coder"}.issubset(set(PUBLIC_SKILLS)))
        self.assertEqual(verify(self.home)["decision"],"pass")

    def test_loopback_api_health_and_auth(self):
        token="T"*32; server=ThreadingHTTPServer(("127.0.0.1",0),_handler(self.home,token)); thread=threading.Thread(target=server.serve_forever,daemon=True); thread.start()
        try:
            conn=http.client.HTTPConnection("127.0.0.1",server.server_port,timeout=5); conn.request("GET","/health"); self.assertEqual(conn.getresponse().status,200); conn.close()
            conn=http.client.HTTPConnection("127.0.0.1",server.server_port,timeout=5); conn.request("GET","/api/meeting/v1/meetings"); self.assertEqual(conn.getresponse().status,401); conn.close()
            conn=http.client.HTTPConnection("127.0.0.1",server.server_port,timeout=5); conn.request("GET","/api/meeting/v1/meetings",headers={"Authorization":f"Bearer {token}"}); response=conn.getresponse(); self.assertEqual(response.status,200); self.assertIn("meetings",json.loads(response.read())); conn.close()
        finally:
            server.shutdown(); server.server_close(); thread.join(timeout=5)

if __name__=="__main__": unittest.main()
