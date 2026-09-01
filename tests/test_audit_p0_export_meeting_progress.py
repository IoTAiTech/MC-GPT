# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Required Notice: Copyright 2026 IoT-AI.Tech / Dr.-Ing. Babak Sorkhpour
# Author: Dr.-Ing. Babak Sorkhpour, with AI assistance
# Version: 6.7.0-beta.5 | Date: 2026-08-14
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from iot_ai.export_gate import assert_export_safe, inspect_export_file
from iot_ai.meeting import _requested_seats, show as show_meeting, start
from iot_ai.projection import export_workspace
from iot_ai.tasks import create, record_progress, show as show_task
from iot_ai.workspace import connect_write
from tests.common import IsolatedHomeTestCase, synthetic_rfc1918_host


class ExportGateFailClosedTests(unittest.TestCase):
    def test_unknown_binary_is_blocked(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "blob.bin"
            path.write_bytes(b"\x00\x01\xff\xfe")
            result = assert_export_safe(path, public=True, allowed_roots=[Path(tmp)])
            self.assertEqual(result["decision"], "block")
            self.assertIn("unknown_binary", result["findings"])

    def test_xlsx_with_private_ip_is_blocked_for_public(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sheet.xlsx"
            wb = Workbook()
            wb.active["A1"] = "host " + synthetic_rfc1918_host()
            wb.save(path)
            result = assert_export_safe(path, public=True, allowed_roots=[Path(tmp)])
            self.assertEqual(result["decision"], "block")
            self.assertIn("private_ip", result.get("findings") or [])

    def test_xlsx_formula_text_is_scanned(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "formula.xlsx"
            wb = Workbook()
            wb.active["A1"] = '="' + synthetic_rfc1918_host() + '"'
            wb.save(path)
            result = assert_export_safe(path, public=True, allowed_roots=[Path(tmp)])
            self.assertEqual(result["decision"], "block")
            self.assertIn("private_ip", result.get("findings") or [])


class ProjectionTaskScopeTests(IsolatedHomeTestCase):
    def test_task_scoped_export_omits_other_tasks(self) -> None:
        first = create(self.home, "First task", "one", "normal", None)
        create(self.home, "Second task", "two", "normal", None)
        output = self.home / "one.xlsx"
        export_workspace(self.home, output, task_id=first["task_id"])
        names = [row[0].value for row in load_workbook(output)["Tasks"].iter_rows(min_row=2, max_col=1) if row[0].value]
        self.assertEqual(names, [first["task_id"]])


class MeetingIsolationTests(IsolatedHomeTestCase):
    def test_requested_seats_are_bound_to_meeting_id(self) -> None:
        first = start(self.home, "First meeting", ["claude"], quorum=1)
        second = start(self.home, "Second meeting", ["codex"], quorum=1, existing_task_id=first["task_id"])
        self.assertEqual(first["task_id"], second["task_id"])
        self.assertNotEqual(first["meeting_id"], second["meeting_id"])
        self.assertEqual(_requested_seats(self.home, first["meeting_id"]), ["claude"])
        self.assertEqual(_requested_seats(self.home, second["meeting_id"]), ["codex"])
        shown = show_meeting(self.home, second["meeting_id"])
        requested = list((shown.get("seat_plan") or {}).get("requested_seats") or [])
        self.assertEqual(requested, ["codex"])


class ProgressTelemetryTests(IsolatedHomeTestCase):
    def test_progress_does_not_force_backlog_or_needs_work_to_active(self) -> None:
        backlog_id = create(self.home, "Backlog task", "desc", "normal", None)["task_id"]
        self.assertEqual(show_task(self.home, backlog_id)["task"]["status"], "backlog")
        self.assertEqual(record_progress(self.home, backlog_id, "observe", 10, "note")["status"], "backlog")
        self.assertEqual(show_task(self.home, backlog_id)["task"]["status"], "backlog")

        needs_id = create(self.home, "Needs work task", "desc", "normal", None)["task_id"]
        conn = connect_write(self.home)
        conn.execute("UPDATE tasks SET status='needs-work' WHERE id=?", (needs_id,))
        conn.commit()
        conn.close()
        self.assertEqual(record_progress(self.home, needs_id, "observe", 20, "still blocked")["status"], "needs-work")
        self.assertEqual(show_task(self.home, needs_id)["task"]["status"], "needs-work")


if __name__ == "__main__":
    unittest.main()
