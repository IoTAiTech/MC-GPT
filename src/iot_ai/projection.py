# SPDX-License-Identifier: LicenseRef-PolyForm-Noncommercial-1.0.0
# Required Notice: Copyright 2026 IoT-AI.Tech / Dr.-Ing. Babak Sorkhpour
# Author: Dr.-Ing. Babak Sorkhpour, with AI assistance
# Version: 6.7.0-beta.5 | Date: 2026-08-08
"""Atomic Excel redundancy/projection for the standalone workspace."""
from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

from .paths import data_root, db_path
from .util import atomic_json, confined_bytes_write, confined_makedirs, resolve_within_allowed_roots, sha256_file, trusted_operator_roots, utc_now
from .workspace import connect_read, connect_write, excel_manifest_path, excel_path, new_id, rows

SHEETS: list[tuple[str, str, list[str]]] = [
    ("Tasks", "SELECT * FROM tasks ORDER BY created_at", ["id","source","source_id","owner","title","priority","risk_class","task_type","status","engineering_stage","engineering_progress","task_progress","revision","blocker","final_decision","result_summary","created_at","updated_at"]),
    ("Work Units", "SELECT * FROM work_units ORDER BY created_at", ["id","task_id","title","role","status","provider","model_requested","model_served","engineering_stage","engineering_progress","revision","created_at","updated_at"]),
    ("Task Validations", "SELECT * FROM task_validations ORDER BY created_at", ["id","task_id","source_revision","applied_revision","trigger_action","policy","status","validation_task_id","validation_meeting_id","plan_digest","verdict","confidence","user_decision","decision_subject","decision_note","created_at","updated_at"]),
    ("Leases", "SELECT id,work_unit_id,task_id,owner,session_id,status,issued_at,heartbeat_at,expires_at,released_at,revision FROM leases ORDER BY issued_at", ["id","work_unit_id","task_id","owner","session_id","status","issued_at","heartbeat_at","expires_at","released_at","revision"]),
    ("Progress", "SELECT * FROM progress_events ORDER BY created_at", ["id","task_id","work_unit_id","stage","percent","summary","created_at"]),
    ("Attempts", "SELECT * FROM attempts ORDER BY created_at", ["id","task_id","work_unit_id","run_id","provider","role","stage","status","model_requested","model_served","request_or_job_id","auth_route","input_tokens","cached_tokens","output_tokens","reasoning_tokens","latency_ms","fallback_used","failure_class","retry_after","created_at"]),
    ("Evidence", "SELECT * FROM evidence ORDER BY created_at", ["id","task_id","work_unit_id","kind","artifact_path","artifact_sha256","exit_code","passed","created_at"]),
    ("Meetings", "SELECT * FROM meetings ORDER BY created_at", ["id","task_id","topic","depth","effort","status","requested_seats","substantive_seats","quorum","rounds","final_decision","user_approved","consultation_sha256","created_at","updated_at"]),
    ("Meeting Seats", "SELECT * FROM meeting_contributions ORDER BY created_at", ["id","meeting_id","task_id","seat","kind","round_no","status","model_requested","model_served","request_or_job_id","auth_route","input_tokens","cached_tokens","output_tokens","reasoning_tokens","latency_ms","fallback_used","failure_class","quality_score","quality_basis","created_at"]),
    ("KPIs", "SELECT * FROM meeting_kpis ORDER BY meeting_id,name", ["id","meeting_id","name","target","measurement","mandatory","created_at"]),
    ("Cases", "SELECT * FROM meeting_cases ORDER BY meeting_id,case_type,ordinal", ["id","meeting_id","case_type","ordinal","title","description","expected","mandatory","created_at"]),
    ("Tests", "SELECT * FROM test_results ORDER BY created_at", ["id","task_id","work_unit_id","run_id","tier","command_sha256","exit_code","passed","failed","skipped","duration_ms","decision","output_sha256","created_at"]),
    ("Audits", "SELECT * FROM audits ORDER BY created_at", ["id","task_id","decision","gate_score","evidence_sha256","created_at"]),
    ("Contributions", "SELECT * FROM contributions ORDER BY created_at", ["id","run_id","task_id","meeting_id","provider","role","stage","model_requested","model_served","request_id","auth_route","input_tokens","cached_tokens","output_tokens","reasoning_tokens","latency_ms","retries","timeout","status","failure_class","retry_after","fallback_used","quality_score","quality_basis","accepted_findings","rejected_findings","created_at"]),
    ("Knowledge", "SELECT * FROM knowledge_items ORDER BY created_at", ["id","source_type","source_id","task_id","title","content_sha256","classification","tags_json","token_estimate","created_at"]),
    ("Graph Runs", "SELECT * FROM graph_runs ORDER BY created_at", ["id","correlation_id","goal","risk_class","privacy_class","status","plan_digest","token_budget","tokens_used","wall_clock_seconds","elapsed_ms","max_parallel","parallel_efficiency","created_at","updated_at"]),
    ("Graph Nodes", "SELECT * FROM graph_nodes ORDER BY created_at", ["id","graph_id","role_id","node_type","stage","required","status","provider","model_requested","model_served","effort_requested","effort_effective","latency_ms","input_tokens","cached_tokens","output_tokens","reasoning_tokens","output_sha256","failure_class","created_at","updated_at"]),
    ("Role Bindings", "SELECT * FROM role_bindings ORDER BY created_at", ["id","graph_id","node_id","role_id","contract_sha256","provider_candidate_id","provider","model","authority_json","output_schema_json","created_at"]),
    ("Knowledge Artifacts", "SELECT * FROM knowledge_artifact_receipts ORDER BY created_at", ["id","artifact_id","kind","visibility","privacy_class","content_sha256","file_path","source_json","status","created_at"]),
    ("Events", "SELECT seq,event_id,task_id,work_unit_id,event_type,prev_hash,event_hash,created_at FROM events ORDER BY seq", ["seq","event_id","task_id","work_unit_id","event_type","prev_hash","event_hash","created_at"]),
]


def _safe(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _scope_clause(sheet_name: str, headers: list[str]) -> str | None:
    if sheet_name == "Tasks":
        return "id=?"
    if "task_id" in headers:
        return "task_id=?"
    if "meeting_id" in headers:
        return "meeting_id IN (SELECT id FROM meetings WHERE task_id=?)"
    return None


def _scoped_query(sheet_name: str, query: str, headers: list[str], task_id: str | None) -> tuple[str, tuple[Any, ...]]:
    if not task_id:
        return query, ()
    if " ORDER BY " not in query:
        raise ValueError("projection query must include ORDER BY")
    body, order = query.split(" ORDER BY ", 1)
    clause = _scope_clause(sheet_name, headers)
    if clause is None:
        return f"{body} WHERE 1=0 ORDER BY {order}", ()
    joiner = " AND " if " WHERE " in body else " WHERE "
    return f"{body}{joiner}{clause} ORDER BY {order}", (task_id,)


def export_workspace(user_home: Path, output: Path | None = None, task_id: str | None = None) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl>=3.1 is required for Excel export") from exc

    output = output or excel_path(user_home)
    roots = list(trusted_operator_roots(user_home))
    confined_makedirs(data_root(user_home), roots)
    output = resolve_within_allowed_roots(output, roots, must_exist=False)
    job_id = new_id("proj")
    now = utc_now()
    write = connect_write(user_home)
    write.execute("INSERT INTO projection_jobs(id,task_id,status,attempts,created_at,updated_at) VALUES(?,?,?,?,?,?)", (job_id, task_id, "running", 1, now, now))
    write.commit(); write.close()

    conn = connect_read(user_home)
    if conn is None:
        raise RuntimeError("workspace database does not exist")
    counts: dict[str, int] = {}
    workbook = Workbook()
    workbook.remove(workbook.active)
    try:
        for sheet_name, query, headers in SHEETS:
            ws = workbook.create_sheet(sheet_name)
            ws.append(headers)
            scoped, params = _scoped_query(sheet_name, query, headers, task_id)
            records = rows(conn, scoped, params)
            counts[sheet_name.lower().replace(" ", "_")] = len(records)
            for record in records:
                ws.append([_safe(record.get(name)) for name in headers])
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(records)+1)}"
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="0E7490")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            for column in ws.columns:
                values = [len(str(cell.value or "")) for cell in column[:200]]
                ws.column_dimensions[get_column_letter(column[0].column)].width = min(48, max(12, max(values, default=10) + 2))
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        buf = io.BytesIO()
        workbook.save(buf)
        confined_bytes_write(output, buf.getvalue(), roots)
        digest = sha256_file(output, allowed_roots=roots, max_bytes=None)
        manifest = {
            "schema": "iot-ai.excel-projection.v3",
            "generated_at": utc_now(),
            "database": str(db_path(user_home)),
            "output": str(output),
            "sha256": digest,
            "counts": counts,
            "canonical": False,
            "notice": "SQLite is canonical. This workbook is a sealed redundant human projection.",
        }
        atomic_json(excel_manifest_path(user_home), manifest)
        write = connect_write(user_home)
        write.execute("UPDATE projection_jobs SET status='pass',output_path=?,output_sha256=?,manifest_path=?,updated_at=? WHERE id=?", (str(output), digest, str(excel_manifest_path(user_home)), utc_now(), job_id))
        write.commit(); write.close()
        return {"decision": "pass", "output": str(output), "sha256": digest, "manifest": str(excel_manifest_path(user_home)), **counts}
    except Exception as exc:
        write = connect_write(user_home)
        write.execute("UPDATE projection_jobs SET status='failed',error=?,updated_at=? WHERE id=?", (type(exc).__name__, utc_now(), job_id))
        write.commit(); write.close()
        raise
    finally:
        conn.close()
